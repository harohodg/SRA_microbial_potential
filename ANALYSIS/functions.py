import os
import pandas as pd
import duckdb
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl import load_workbook
from IPython.display import display, HTML
#display(HTML("<style>:root { --jp-notebook-max-width: 100% !important; }</style>"))

PLOT_FILE_TYPES = ['html', 'svg', 'png', 'pdf']

import os
from pathlib import Path

def run_sql_query(sql, database = '', parameters={}, show_progress=True):
    if '\n' not in sql and Path(sql).is_file():
        with open(sql) as f:
            query = f.read()
    else:
        query = sql
    query = query.format( **parameters ) if parameters != {} else query

    query = "SET enable_progress_bar = true;" + query if show_progress else query
    
    if database == '':
        results = duckdb.sql( query ).df()
    else:
        with duckdb.connect(database, config = {'access_mode' : 'READ_ONLY'}) as con:
            data_table = con.sql( query ).df()
        results = data_table
    return query, results

def ensure_parent_folder(file_path):
    # Get the parent directory of the file
    parent_dir = os.path.dirname(file_path)
    
    # Check if the parent directory exists
    if not os.path.exists(parent_dir):
        os.makedirs(parent_dir)

def display_table(data_table, with_index = False):
    return HTML( data_table.to_html(index=with_index) )

def export_table(data_table, fname, separator='\t', with_index=False, with_header=True):
    ensure_parent_folder(fname)
    data_table.to_csv(fname, sep=separator, encoding='utf-8', index=with_index, header=with_header)

def export_plotly_plot(figure, plot_name, file_endings=PLOT_FILE_TYPES):
    ensure_parent_folder(plot_name)
    for ending in PLOT_FILE_TYPES:
        fname = f'{plot_name}.{ending}'
        figure.write_html(fname) if ending == 'html' else figure.write_image(fname)

def export_altair_plot(chart, plot_name, file_endings=PLOT_FILE_TYPES, ppi=300):
    ensure_parent_folder(plot_name)
    for ending in PLOT_FILE_TYPES:
        fname = f'{plot_name}.{ending}'
        chart.save(fname, ppi=ppi) if ending == 'png' else chart.save(fname)

def export_plotly_plot(fig, plot_name, plot_config={}, file_endings=PLOT_FILE_TYPES):
    ensure_parent_folder(plot_name)
    for ending in PLOT_FILE_TYPES:
        fname = f'{plot_name}.{ending}'
        if ending == 'html':
            fig.write_html(fname, config=plot_config)
        else:
            plot_width  = plot_config.get('width', None)
            plot_height = plot_config.get('height', None)
            fig.write_image(fname, width=plot_width, height=plot_height)

def export_ggplot_plot(plot, plot_fname, plot_config={}, file_endings=['svg', 'png']):
    for ending in file_endings:
        plot.save(f"{plot_fname}.{ending}", **plot_config) 


def export_excell_sheet(fname, data_table, sheet = 'Sheet 1', to_merge = [], header=[], header_style={'font' : Font(bold=True)}):
    ensure_parent_folder(fname)
    parameters = {'mode' : 'a' if Path(fname).is_file() else 'w'}
    if parameters['mode'] == 'a':
        parameters['if_sheet_exists'] = 'replace'
        
    with pd.ExcelWriter(fname, **parameters) as writer:
        data_table.to_excel(writer, sheet_name=sheet, startrow=len(header) , startcol=0, index=False)   

    if len(header) != 0 or len(to_merge) != 0:
        workbook = load_workbook(filename=fname)
        sheet = workbook[sheet]
        
        for index, string in enumerate(header):
            sheet[f"A{index+1}"] = string
            for key, value in header_style.items():
                setattr(sheet[f"A{index+1}"], key, value)  
        
        for first_row, last_row, first_column, last_column in to_merge:
            sheet.merge_cells(start_row=first_row, start_column=first_column, end_row=last_row, end_column=last_column)  
        
        workbook.save(filename=fname)